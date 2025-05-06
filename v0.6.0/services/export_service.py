"""
Сервис для экспорта данных.
"""
import io
import csv
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from typing import Optional, List, Dict, Any
from sqlalchemy.orm import Session
from sqlalchemy import desc

from models import Order, Client, Service, Employee, Expense, Payment
from services.order_service import OrderService as OrderSvc
from services.employee_service import EmployeeService
from services.finance_service import FinanceService

class ExportService:
    """
    Сервис для экспорта данных в различные форматы.
    """
    
    @staticmethod
    def export_orders(
        db: Session,
        format: str = "csv",
        status: Optional[str] = None,
        client_name: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None
    ):
        """
        Экспорт данных о заказах с фильтрацией.
        """
        # Формируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_orders_{timestamp}.{format}"
        
        # Получаем заказы с фильтрацией
        orders_data = OrderSvc.get_orders(
            db=db,
            status=status,
            client_name=client_name,
            date_from=date_from,
            date_to=date_to,
            page=1,
            limit=1000  # Для экспорта берем максимум 1000 заказов
        )
        
        orders = orders_data["orders"]
        
        if format == "csv":
            # Экспорт в CSV
            output = io.StringIO()
            writer = csv.writer(output, lineterminator='\n')
            
            # Заголовки
            headers = [
                "ID", "Клиент", "Телефон", "Менеджер", "Монтажники",
                "Дата заказа", "Дата завершения", "Статус", "Стоимость монтажа (₽)",
                "Услуги", "Общая сумма (₽)", "Примечания"
            ]
            writer.writerow(headers)
            
            # Данные
            for order in orders:
                # Формируем строку монтажников
                installers = ", ".join([emp["name"] for emp in order["employees"] if emp["employee_type"] == "монтажник"])
                
                # Формируем строку услуг
                services = ", ".join([svc["name"] for svc in order["services"]])
                
                row = [
                    order["id"],
                    order["client"]["name"] if order["client"] else "-",
                    order["client"]["phone"] if order["client"] else "-",
                    order["manager"]["name"] if order["manager"] else "-",
                    installers,
                    order["order_date"],
                    order["completion_date"] if order["completion_date"] else "-",
                    order["status"],
                    order["mount_price"],
                    services,
                    order["total_price"],
                    order["notes"] if order["notes"] else "-"
                ]
                writer.writerow(row)
            
            return {"data": output.getvalue(), "filename": filename, "media_type": "text/csv"}
        
        elif format == "xlsx":
            # Экспорт в Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Заказы"
            
            # Заголовки
            headers = [
                "ID", "Клиент", "Телефон", "Менеджер", "Монтажники",
                "Дата заказа", "Дата завершения", "Статус", "Стоимость монтажа (₽)",
                "Услуги", "Общая сумма (₽)", "Примечания"
            ]
            sheet.append(headers)
            
            # Выделяем заголовки жирным
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            
            # Данные
            for order in orders:
                # Формируем строку монтажников
                installers = ", ".join([emp["name"] for emp in order["employees"] if emp["employee_type"] == "монтажник"])
                
                # Формируем строку услуг
                services = ", ".join([svc["name"] for svc in order["services"]])
                
                row = [
                    order["id"],
                    order["client"]["name"] if order["client"] else "-",
                    order["client"]["phone"] if order["client"] else "-",
                    order["manager"]["name"] if order["manager"] else "-",
                    installers,
                    order["order_date"],
                    order["completion_date"] if order["completion_date"] else "-",
                    order["status"],
                    order["mount_price"],
                    services,
                    order["total_price"],
                    order["notes"] if order["notes"] else "-"
                ]
                sheet.append(row)
            
            # Автоматическая ширина столбцов
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем в буфер
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return {"data": output.getvalue(), "filename": filename, "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        
        else:
            return {"error": "Неподдерживаемый формат экспорта. Поддерживаются: csv, xlsx"}
    
    @staticmethod
    def export_clients(
        db: Session,
        format: str = "csv",
        search: Optional[str] = None,
        source: Optional[str] = None
    ):
        """
        Экспорт данных о клиентах с фильтрацией.
        """
        # Формируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_clients_{timestamp}.{format}"
        
        # Базовый запрос
        query = db.query(Client)
        
        # Фильтрация
        if search:
            from sqlalchemy import or_
            query = query.filter(
                or_(
                    Client.name.ilike(f"%{search}%"),
                    Client.phone.ilike(f"%{search}%")
                )
            )
        
        if source and source != "Все":
            query = query.filter(Client.source == source)
        
        # Сортировка
        query = query.order_by(Client.name)
        
        # Получаем данные
        clients = query.all()
        
        if format == "csv":
            # Экспорт в CSV
            output = io.StringIO()
            writer = csv.writer(output, lineterminator='\n')
            
            # Заголовки
            headers = ["ID", "Имя", "Телефон", "Источник", "Дата регистрации", "Дата обновления", "Количество заказов"]
            writer.writerow(headers)
            
            # Данные
            for client in clients:
                # Подсчет заказов клиента
                order_count = db.query(Order).filter(Order.client_id == client.id).count()
                
                row = [
                    client.id,
                    client.name,
                    client.phone,
                    client.source,
                    client.created_at,
                    client.updated_at if client.updated_at else "-",
                    order_count
                ]
                writer.writerow(row)
            
            return {"data": output.getvalue(), "filename": filename, "media_type": "text/csv"}
        
        elif format == "xlsx":
            # Экспорт в Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Клиенты"
            
            # Заголовки
            headers = ["ID", "Имя", "Телефон", "Источник", "Дата регистрации", "Дата обновления", "Количество заказов"]
            sheet.append(headers)
            
            # Выделяем заголовки жирным
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            
            # Данные
            for client in clients:
                # Подсчет заказов клиента
                order_count = db.query(Order).filter(Order.client_id == client.id).count()
                
                row = [
                    client.id,
                    client.name,
                    client.phone,
                    client.source,
                    client.created_at,
                    client.updated_at if client.updated_at else "-",
                    order_count
                ]
                sheet.append(row)
            
            # Автоматическая ширина столбцов
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем в буфер
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return {"data": output.getvalue(), "filename": filename, "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        
        else:
            return {"error": "Неподдерживаемый формат экспорта. Поддерживаются: csv, xlsx"}
    
    @staticmethod
    def export_services(
        db: Session,
        format: str = "csv",
        search: Optional[str] = None,
        category: Optional[str] = None
    ):
        """
        Экспорт данных об услугах с фильтрацией.
        """
        # Формируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_services_{timestamp}.{format}"
        
        # Базовый запрос
        query = db.query(Service)
        
        # Фильтрация
        if search:
            query = query.filter(Service.name.ilike(f"%{search}%"))
        
        if category and category != "Все":
            query = query.filter(Service.category == category)
        
        # Сортировка
        query = query.order_by(Service.name)
        
        # Получаем данные
        services = query.all()
        
        if format == "csv":
            # Экспорт в CSV
            output = io.StringIO()
            writer = csv.writer(output, lineterminator='\n')
            
            # Заголовки
            headers = [
                "ID", "Название", "Категория", "Закупочная цена (₽)", "Продажная цена (₽)",
                "Базовая цена", "Бонус менеджеру", "Бонус монтажнику (₽)", "Процент прибыли",
                "Дата создания", "Дата обновления"
            ]
            writer.writerow(headers)
            
            # Данные
            for service in services:
                row = [
                    service.id,
                    service.name,
                    service.category,
                    service.purchase_price,
                    service.selling_price,
                    service.default_price if service.default_price else "-",
                    "Да" if service.is_manager_bonus else "Нет",
                    service.installer_bonus_fixed,
                    f"{service.profit_margin_percent * 100:.0f}%",
                    service.created_at,
                    service.updated_at if service.updated_at else "-"
                ]
                writer.writerow(row)
            
            return {"data": output.getvalue(), "filename": filename, "media_type": "text/csv"}
        
        elif format == "xlsx":
            # Экспорт в Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Услуги"
            
            # Заголовки
            headers = [
                "ID", "Название", "Категория", "Закупочная цена (₽)", "Продажная цена (₽)",
                "Базовая цена", "Бонус менеджеру", "Бонус монтажнику (₽)", "Процент прибыли",
                "Дата создания", "Дата обновления"
            ]
            sheet.append(headers)
            
            # Выделяем заголовки жирным
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            
            # Данные
            for service in services:
                row = [
                    service.id,
                    service.name,
                    service.category,
                    service.purchase_price,
                    service.selling_price,
                    service.default_price if service.default_price else "-",
                    "Да" if service.is_manager_bonus else "Нет",
                    service.installer_bonus_fixed,
                    f"{service.profit_margin_percent * 100:.0f}%",
                    service.created_at,
                    service.updated_at if service.updated_at else "-"
                ]
                sheet.append(row)
            
            # Автоматическая ширина столбцов
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем в буфер
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return {"data": output.getvalue(), "filename": filename, "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        
        else:
            return {"error": "Неподдерживаемый формат экспорта. Поддерживаются: csv, xlsx"}
    
    @staticmethod
    def export_employees(
        db: Session,
        format: str = "csv",
        employee_type: Optional[str] = None,
        active: Optional[int] = None,
        month: Optional[str] = None
    ):
        """
        Экспорт данных о сотрудниках с фильтрацией.
        """
        # Формируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_employees_{timestamp}.{format}"
        
        # Если месяц не указан, используем текущий
        if not month:
            month = datetime.now().strftime("%Y-%m")
        
        # Базовый запрос
        query = db.query(Employee)
        
        # Фильтрация
        if employee_type:
            query = query.filter(Employee.employee_type == employee_type)
        
        if active is not None:
            query = query.filter(Employee.active == active)
        
        # Сортировка
        query = query.order_by(Employee.name)
        
        # Получаем данные
        employees = query.all()
        
        # Рассчитываем зарплаты
        employees_with_salary = []
        for employee in employees:
            # Расчет зарплаты
            salary_data = EmployeeService.calculate_salary(db, employee.id, month)
            
            # Если не удалось рассчитать зарплату, пропускаем
            if not salary_data:
                continue
            
            employees_with_salary.append({
                "id": employee.id,
                "name": employee.name,
                "phone": employee.phone,
                "employee_type": employee.employee_type,
                "base_salary": employee.base_salary or 0,
                "active": employee.active,
                "created_at": employee.created_at,
                "updated_at": employee.updated_at,
                "salary": salary_data["salary"],
                "paid": salary_data["paid"],
                "to_pay": salary_data["to_pay"],
                "order_count": len(salary_data["details"]["breakdown"]["orders"])
            })
        
        if format == "csv":
            # Экспорт в CSV
            output = io.StringIO()
            writer = csv.writer(output, lineterminator='\n')
            
            # Заголовки
            headers = [
                "ID", "Имя", "Телефон", "Тип сотрудника", "Базовая зарплата (₽)",
                "Активен", "Заработано (₽)", "Выплачено (₽)", "К выплате (₽)",
                "Количество заказов", "Дата создания", "Дата обновления"
            ]
            writer.writerow(headers)
            
            # Данные
            for employee in employees_with_salary:
                row = [
                    employee["id"],
                    employee["name"],
                    employee["phone"],
                    employee["employee_type"],
                    employee["base_salary"],
                    "Да" if employee["active"] == 1 else "Нет",
                    employee["salary"],
                    employee["paid"],
                    employee["to_pay"],
                    employee["order_count"],
                    employee["created_at"],
                    employee["updated_at"] if employee["updated_at"] else "-"
                ]
                writer.writerow(row)
            
            return {"data": output.getvalue(), "filename": filename, "media_type": "text/csv"}
        
        elif format == "xlsx":
            # Экспорт в Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Сотрудники"
            
            # Заголовки
            headers = [
                "ID", "Имя", "Телефон", "Тип сотрудника", "Базовая зарплата (₽)",
                "Активен", "Заработано (₽)", "Выплачено (₽)", "К выплате (₽)",
                "Количество заказов", "Дата создания", "Дата обновления"
            ]
            sheet.append(headers)
            
            # Выделяем заголовки жирным
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            
            # Данные
            for employee in employees_with_salary:
                row = [
                    employee["id"],
                    employee["name"],
                    employee["phone"],
                    employee["employee_type"],
                    employee["base_salary"],
                    "Да" if employee["active"] == 1 else "Нет",
                    employee["salary"],
                    employee["paid"],
                    employee["to_pay"],
                    employee["order_count"],
                    employee["created_at"],
                    employee["updated_at"] if employee["updated_at"] else "-"
                ]
                sheet.append(row)
            
            # Автоматическая ширина столбцов
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем в буфер
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return {"data": output.getvalue(), "filename": filename, "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        
        else:
            return {"error": "Неподдерживаемый формат экспорта. Поддерживаются: csv, xlsx"}
    
    @staticmethod
    def export_finances(
        db: Session,
        format: str = "csv",
        date_from: Optional[str] = None,
        date_to: Optional[str] = None
    ):
        """
        Экспорт финансовых данных с фильтрацией по датам.
        """
        # Формируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_finances_{timestamp}.{format}"
        
        # Получаем финансовую сводку
        summary = FinanceService.get_finance_summary(db, date_from, date_to)
        
        # Получаем историю транзакций
        transactions_data = FinanceService.get_transaction_history(
            db=db,
            date_from=date_from,
            date_to=date_to,
            page=1,
            limit=1000  # Для экспорта берем максимум 1000 транзакций
        )
        
        transactions = transactions_data["transactions"]
        
        if format == "csv":
            # Экспорт в CSV
            output = io.StringIO()
            writer = csv.writer(output, lineterminator='\n')
            
            # Заголовки для сводки
            headers = ["Период", "", ""]
            writer.writerow(headers)
            
            if date_from and date_to:
                period = f"с {date_from} по {date_to}"
            elif date_from:
                period = f"с {date_from}"
            elif date_to:
                period = f"по {date_to}"
            else:
                period = "весь период"
            
            writer.writerow([period, "", ""])
            writer.writerow(["", "", ""])
            
            # Данные сводки
            writer.writerow(["Выручка (₽)", summary["total_revenue"], ""])
            writer.writerow(["Расходы (₽)", summary["total_expenses"], ""])
            writer.writerow(["Комиссии (₽)", summary["total_commissions"], ""])
            writer.writerow(["Прибыль (₽)", summary["total_profit"], ""])
            writer.writerow(["Текущий баланс (₽)", summary["current_balance"], ""])
            writer.writerow(["", "", ""])
            
            # Заголовки для расходов по категориям
            writer.writerow(["Расходы по категориям", "", ""])
            for category, amount in summary["expenses_by_category"].items():
                writer.writerow([category, amount, ""])
            
            writer.writerow(["", "", ""])
            
            # Заголовки для доходов по источникам
            writer.writerow(["Доходы по источникам", "", ""])
            for source, amount in summary["revenue_by_source"].items():
                writer.writerow([source, amount, ""])
            
            writer.writerow(["", "", ""])
            
            # Заголовки для транзакций
            writer.writerow(["История транзакций", "", ""])
            trans_headers = ["ID", "Дата", "Тип", "Источник", "Сумма (₽)", "Описание"]
            writer.writerow(trans_headers)
            
            # Данные транзакций
            for transaction in transactions:
                row = [
                    transaction.id,
                    transaction.transaction_date,
                    transaction.transaction_type,
                    transaction.source_type,
                    transaction.amount,
                    transaction.description or "-"
                ]
                writer.writerow(row)
            
            return {"data": output.getvalue(), "filename": filename, "media_type": "text/csv"}
        
        elif format == "xlsx":
            # Экспорт в Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Финансы"
            
            # Заголовки для сводки
            headers = ["Период", "", ""]
            sheet.append(headers)
            
            if date_from and date_to:
                period = f"с {date_from} по {date_to}"
            elif date_from:
                period = f"с {date_from}"
            elif date_to:
                period = f"по {date_to}"
            else:
                period = "весь период"
            
            sheet.append([period, "", ""])
            sheet.append(["", "", ""])
            
            # Данные сводки
            sheet.append(["Выручка (₽)", summary["total_revenue"], ""])
            sheet.append(["Расходы (₽)", summary["total_expenses"], ""])
            sheet.append(["Комиссии (₽)", summary["total_commissions"], ""])
            sheet.append(["Прибыль (₽)", summary["total_profit"], ""])
            sheet.append(["Текущий баланс (₽)", summary["current_balance"], ""])
            sheet.append(["", "", ""])
            
            # Заголовки для расходов по категориям
            sheet.append(["Расходы по категориям", "", ""])
            for category, amount in summary["expenses_by_category"].items():
                sheet.append([category, amount, ""])
            
            sheet.append(["", "", ""])
            
            # Заголовки для доходов по источникам
            sheet.append(["Доходы по источникам", "", ""])
            for source, amount in summary["revenue_by_source"].items():
                sheet.append([source, amount, ""])
            
            sheet.append(["", "", ""])
            
            # Создаем второй лист для транзакций
            trans_sheet = workbook.create_sheet(title="Транзакции")
            
            # Заголовки для транзакций
            trans_headers = ["ID", "Дата", "Тип", "Источник", "Сумма (₽)", "Описание"]
            trans_sheet.append(trans_headers)
            
            # Выделяем заголовки жирным
            for cell in trans_sheet[1]:
                cell.font = Font(bold=True)
            
            # Данные транзакций
            for transaction in transactions:
                row = [
                    transaction.id,
                    transaction.transaction_date,
                    transaction.transaction_type,
                    transaction.source_type,
                    transaction.amount,
                    transaction.description or "-"
                ]
                trans_sheet.append(row)
            
            # Автоматическая ширина столбцов для обоих листов
            for sheet_name in workbook.sheetnames:
                current_sheet = workbook[sheet_name]
                for column in current_sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                    adjusted_width = (max_length + 2)
                    current_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем в буфер
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return {"data": output.getvalue(), "filename": filename, "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        
        else:
            return {"error": "Неподдерживаемый формат экспорта. Поддерживаются: csv, xlsx"}